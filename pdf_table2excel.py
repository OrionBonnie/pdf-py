import os
import subprocess
import csv
from msvcrt import getch
from pdf_convert import get_pdf_list
from time import sleep
print("基本模块已导入")
try:
    check_java = subprocess.run("java -version", capture_output=True)
except FileNotFoundError:
    print("请检查是否安装好 JDK 1.8 且配置好环境变量")
    print("按任意键退出")
    getch()
    exit()
finally:
    print("依赖检查完成, 为可用状态")

try:
    import pypdf
except ModuleNotFoundError:
    list_mod_all = subprocess.run("pip list --disable-pip-version-check", capture_output=True)
    list_mod_all = list_mod_all.stdout.decode("utf-8")
    if list_mod_all.find("pypdf") == -1:
        subprocess.run("pip install pypdf")
        import pypdf
finally:
    print("pdf读取模块已导入")

try:
    import tabula
except ModuleNotFoundError:
    list_mod_all = subprocess.run("pip list --disable-pip-version-check", capture_output=True)
    list_mod_all = list_mod_all.stdout.decode("utf-8")
    if list_mod_all.find("tabula-py") == -1:
        subprocess.run("pip install tabula-py jpype1")
        import tabula
finally:
    print("列表读取模块已导入")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Side
except ModuleNotFoundError:
    list_mod_all = subprocess.run("pip list --disable-pip-version-check", capture_output=True)
    list_mod_all = list_mod_all.stdout.decode("utf-8")
    if list_mod_all.find("openpyxl") == -1:
        subprocess.run("pip install openpyxl")
        from openpyxl import Workbook
finally:
    print("excel格式模块已导入")

try:
    import tqdm
except ModuleNotFoundError:
    list_mod_all = subprocess.run("pip list --disable-pip-version-check", capture_output=True)
    list_mod_all = list_mod_all.stdout.decode("utf-8")
    if list_mod_all.find("tqdm") == -1:
        subprocess.run("pip install tqdm")
        from tqdm import tqdm


def is_empty_row(in_row):        # 用于判断内容的行是否为空，如果为空，则返回True，用于消除表格空行
    result_list = list()
    for test_null in in_row:
        result_list.append(not bool(test_null))
    return all(result_list)             # all()方法如果所有都为True，则返回True


def choose_page(pdf_path):
    pdf_pages_total = len(pypdf.PdfReader(pdf_path).pages)
    print('''
        当前正在对 {} 操作(共 {} 页):
        1. 全部提取
        2. 指定页数范围(默认从第1页开始, 指定范围时用空格把两个页数隔开, 如: 3 6)
        3. 指定具体页数(用空格把多个页数隔开, 如: 1 3 6 7)
        '''.format(pdf_path, pdf_pages_total))
    choice = input("请输入指定的选项数字(1,2,3): ")
    selected_pages = list()
    if choice == "1":
        for i in range(pdf_pages_total):
            selected_pages.append(i+1)
    if choice == "2":
        page_between = input("请输入你想要提取的范围: ")
        split = page_between.split()
        if len(split) != 2:
            print(selected_pages)
            print("你只可输入两个数字作为范围")
            print("按任意键退出")
            getch()
            exit()
        for i in range(int(split[0]), int(split[1]) + 1):
            selected_pages.append(i)
    if choice == "3":
        page_between = input("请输入你想要提取的页码: ")
        split = page_between.split()
        selected_pages = split.copy()
    if choice == "2" or choice == "3":
        if int(selected_pages[-1]) > pdf_pages_total:
            print("页数超过PDF文件的最大页数, 您输入了大于PDF页数的页码")
            print("你只可输入两个数字作为范围")
            print("按任意键退出")
            getch()
            exit()
    return selected_pages


def pdf_table2csv(pdf_path):
    selected = choose_page(pdf_path)
    if not os.path.exists(r".\csv"):        # 检查csv目录是否存在，不存在则创建
        os.mkdir(r".\csv")
    csv_file = str()
    if os.path.splitext(pdf_path)[-1] == ".pdf":
        cvs_name = os.path.basename(pdf_path).replace("pdf", "csv")
        csv_file = cvs_name
    elif os.path.splitext(pdf_path)[-1] == ".PDF":
        cvs_name = os.path.basename(pdf_path).replace("PDF", "csv")
        csv_file = cvs_name
    csv_path = os.path.join(r".\csv", csv_file)
    tabula.convert_into(pdf_path, csv_path, output_format="csv", pages=selected, lattice=True)
    print(f"开始提取转换 {pdf_path}")
    return csv_path


def csv2xlsx(csv_file):

    rows_content = list()
    csv_io = open(csv_file, "r", encoding="utf-8")
    csv_reader = csv.reader(csv_io)
    for csv_row in csv_reader:
        if not is_empty_row(csv_row):
            rows_content.append(csv_row)

    web_book = Workbook()
    sheet = web_book.active
    row_num = 1      # 当前所在的行
    column_max_num = 1
    for row in tqdm(rows_content):            # 开始遍历行内容列表
        for i in range(len(row)):       # 遍历当前行的所有元素, 此处用于修改当前行单元格的样式
            i += 1
            border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                            bottom=Side(style="thin"))
            sheet.cell(row=row_num, column=i).border = border
            alignment = Alignment(horizontal="left", vertical="center", wrapText=False, shrinkToFit=True)
            sheet.cell(row=row_num, column=i).alignment = alignment
            if i > column_max_num:
                column_max_num = i
        sheet.append(row)
        row_num += 1        # +1使下次遍历对下一行进行格式化

    init_column_ascii = ord("A")
    for column in range(column_max_num):       # 根据最大列数调整列宽
        sheet.column_dimensions[chr(init_column_ascii)].width = 18
        init_column_ascii += 1

    xlsx_name = os.path.basename(csv_file).replace("csv", "xlsx")
    xlsx_save_path = os.path.join(r".\out", xlsx_name)
    web_book.save(xlsx_save_path)


def main():
    if not os.path.exists(r".\out"):
        os.mkdir(r".\out")
    else:
        for file in os.listdir(r".\out"):
            os.remove(r".\out\\" + file)
    for pdf_file in get_pdf_list():
        csv2xlsx(pdf_table2csv(pdf_file))
    for file in os.listdir(r".\csv"):
        os.remove(r".\csv\\" + file)
    os.rmdir(r".\csv")

    print()
    print('转换已完成，共从 {} 个文件中提取了表格'.format(len(os.listdir(r".\out"))))
    print("脚本为了表格输出整洁默认开启字体适应缩放，若一个单元格字数过多，该单元格内的字体也会变小，双击单元格即可恢复原样")
    sleep(1.5)
    os.system(r"start .\out")
    print("按任意键退出关闭此窗口")
    getch()
